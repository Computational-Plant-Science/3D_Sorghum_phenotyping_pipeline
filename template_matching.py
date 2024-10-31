'''
Name: ai_seg.py

Version: 1.0

Summary: A machine learning model U2net and opencv based color clustering method hat performs object segmentation in a single shot
    
Author: Suxing Liu

Author-email: suxingliu@gmail.com

Created: 2024-09-11

USAGE:

    Default parameters: python3 ai_seg.py -p /input/ -ft jpg  -o /output/

PARAMETERS:
    ("-p", "--path", dest = "path", type = str, required = True,    help = "path to image file")
    ("-ft", "--filetype", dest = "filetype", type = str, required = False, default='jpg,png', help = "Image filetype")
    ("-o", "--output_path", dest = "output_path", type = str, required = False,    help = "result path")
    ('-pl', '--parallel', dest = "parallel", type = int, required = False, default = 0,  help = 'Whether using parallel processing or loop processing, 0: Loop, 1: Parallel')

INPUT:
    Image file in jpg, png format

OUTPUT:

    Segmentation results in masked foreground image


'''




# import the necessary packages
import os
import glob
import shutil 
import pathlib
from pathlib import Path

import numpy as np
import argparse
import cv2
import imutils
import openpyxl
import random as nd
from imutils.object_detection import non_max_suppression

from sklearn.cluster import KMeans


from rembg import remove

import time

MBFACTOR = float(1<<20)



# generate folder to store the output results
def mkdir(path):
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        shutil.rmtree(path)
        os.makedirs(path)
        print ("{} path exists!\n".format(path))
        return False
        



# Detect markers in the image
def marker_detect(img_rgb):
    
    
    # get the dimension of the image
    img_height, img_width, img_channels = img_rgb.shape

    
    # convert the input image to a grayscale
    if img_channels > 2:
    
        img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)
    else:
        img_gray = orig
    
    

    ###################################################################
    
    # Store width and height of template in w and h 
    w, h = template.shape[::-1]
    
    img_overlay = img_rgb.copy()
    

    
    # Passing the image to matchTemplate method 
    match = cv2.matchTemplate(image = img_gray, templ=template, method=cv2.TM_CCOEFF_NORMED)
    
    # Specify a threshold 
    threshold = 0.5

    # Select rectangles with 
    # confidence greater than threshold 
    (y_points, x_points) = np.where(match >= threshold) 

    # initialize our list of rectangles 
    boxes = list() 

    # loop over the starting (x, y)-coordinates again 
    for (x, y) in zip(x_points, y_points): 

        # update our list of rectangles 
        boxes.append((x, y, x + w, y + h))
        


    # apply non-maxima suppression to the rectangles 
    # this will create a single bounding box 
    boxes = non_max_suppression(np.array(boxes)) 
    
    blank_image = np.zeros(img_rgb.shape, np.uint8)
    
    
    # loop over the final bounding boxes 
    for (x1, y1, x2, y2) in boxes: 

        # draw the bounding box on the image 
        #img_overlay = cv2.rectangle(blank_image, (x1, y1), (x2, y2), (0, 255, 255), 3)
        blank_image = cv2.rectangle(blank_image, (x1, y1), (x2, y2), (255, 255, 255), -1) 
    
    
    (ret, thresh_mask) = cv2.threshold(cv2.cvtColor(blank_image, cv2.COLOR_BGR2GRAY), 127, 255, 0)
    
    #mask = np.zeros(image.shape[:2], dtype="uint8")
    
    #cv2.rectangle(mask, (0, 90), (290, 450), 255, -1)
    
    # use mask to generate segmentation object
    masked_marker = cv2.bitwise_and(img_rgb, img_rgb, mask = thresh_mask)
    
    #img_overlay = cv2.bitwise_and(img_rgb, img_rgb, mask = thresh_mask)
    
    
    ####################################################################
    # AI pre-trained model to segment plant object, return mask
    masked_marker_seg = remove(masked_marker, only_mask = True).copy()
    
    #img_overlay = remove(masked_marker, only_mask = True).copy()
    
    #masked_marker_seg_gray = cv2.cvtColor(masked_marker_seg, cv2.COLOR_BGR2GRAY)
    
    
    
    # Apply thresholding to create a binary image
    _, thresh = cv2.threshold(masked_marker_seg, 127, 255, cv2.THRESH_BINARY)

    # Find contours in the binary image
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    print("Number of contours detected:\n", len(contours))
    
    
    width_rec = []
    
    
    # Draw bounding boxes around the contours
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        img_overlay = cv2.rectangle(masked_marker, (x, y), (x + w, y + h), (0, 255, 0), 2)
        
        bb_w = min(w, h)
        
        if bb_w> 300 and bb_w < 400:
        
            print("Bounding box size: {}\n".format((bb_w)))
            
            width_rec.append(bb_w)

    
    ###################################################################
    # compute the average of detected square dimension in pixels
    if len(width_rec) > 0:
        
        avg_width = np.mean(width_rec)
        
        pixel_cm_ratio = avg_width/50.8
     
    else:
        
        avg_width = 0
        pixel_cm_ratio = 0
    
    return img_overlay, avg_width, pixel_cm_ratio
    


 
    


# segment foreground object using color clustering method
def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters):
    

    #image_LAB = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
    
    #cl = ColorLabeler()
    
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        
    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)
        
    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        
    else:
        colorSpace = 'bgr'  # set for file naming purposes

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
            
    (height, width, n_channel) = image.shape
    
    if n_channel > 1:
        
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image
        
 
    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])
    
    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')
    
    # define number of cluster, at lease 2 cluster including background
    numClusters = max(2, args_num_clusters)
    
    # clustering method
    kmeans = KMeans(n_clusters = numClusters, n_init = 40, max_iter = 500).fit(reshaped)
    
    # get lables 
    pred_label = kmeans.labels_
    
    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)],key = lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i
    
    (ret, thresh) = cv2.threshold(kmeansImage,0,255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    '''
    if np.count_nonzero(thresh) > 0:
        
        thresh_cleaned = clear_border(thresh)
    else:
        thresh_cleaned = thresh
    '''
    
    #thresh_cleaned = thresh
    
    img_thresh = thresh
    
    '''
    (numLabels, labels, stats, centroids) = cv2.connectedComponentsWithStats(thresh_cleaned, connectivity = 8)
     
    
    # stats[0], centroids[0] are for the background label. ignore
    # cv2.CC_STAT_LEFT, cv2.CC_STAT_TOP, cv2.CC_STAT_WIDTH, cv2.CC_STAT_HEIGHT
    
    # extract the connected component statistics for the current label
    sizes = stats[1:, cv2.CC_STAT_AREA]
    Coord_left = stats[1:, cv2.CC_STAT_LEFT]
    Coord_top = stats[1:, cv2.CC_STAT_TOP]
    Coord_width = stats[1:, cv2.CC_STAT_WIDTH]
    Coord_height = stats[1:, cv2.CC_STAT_HEIGHT]
    Coord_centroids = np.delete(centroids,(0), axis=0)
    

    
    #print("Coord_centroids {}\n".format(centroids[1][1]))
    
    #print("[width, height] {} {}\n".format(width, height))
    
    numLabels = numLabels - 1
    '''

    
    
    ################################################################################################

    '''
    min_size = 100
    max_size = min(width*height, args_max_size)

    # initialize an output mask
    mask = np.zeros(gray.shape, dtype="uint8")
    
    # loop over the number of unique connected component labels, skipping
    # over the first label (as label zero is the background)
    for i in range(1, numLabels):
    # extract the connected component statistics for the current label
        x = stats[i, cv2.CC_STAT_LEFT]
        y = stats[i, cv2.CC_STAT_TOP]
        w = stats[i, cv2.CC_STAT_WIDTH]
        h = stats[i, cv2.CC_STAT_HEIGHT]
        area = stats[i, cv2.CC_STAT_AREA]
    
        
        # ensure the width, height, and area are all neither too small
        # nor too big
        keepWidth = w > 0 and w < 6000
        keepHeight = h > 0 and h < 4000
        keepArea = area > min_size and area < max_size
        
        #if all((keepWidth, keepHeight, keepArea)):
        # ensure the connected component we are examining passes all three tests
        #if all((keepWidth, keepHeight, keepArea)):
        if keepArea:
        # construct a mask for the current connected component and
        # then take the bitwise OR with the mask
            print("[INFO] keeping connected component '{}'".format(i))
            componentMask = (labels == i).astype("uint8") * 255
            mask = cv2.bitwise_or(mask, componentMask)
            
    
    img_thresh = mask
    
    '''
    ###################################################################################################
    '''
    size_kernel = 5
    
    #if mask contains mutiple non-connected parts, combine them into one. 
    (contours, hier) = cv2.findContours(img_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    if len(contours) > 1:
        
        print("mask contains mutiple non-connected parts, combine them into one\n")
        
        kernel = np.ones((size_kernel,size_kernel), np.uint8)

        dilation = cv2.dilate(img_thresh.copy(), kernel, iterations = 1)
        
        closing = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
        
        img_thresh = closing
    '''
    
    return img_thresh




# compute all the traits
def u2net_seg(image_file):

    ################################################################################
    # load image data

    if args['filetype'] == 'jpg' or args['filetype']:
        
        image = cv2.imread(image_file)
        
    else:
        
        print("Image foramt was not support for now\n")
        sys.exit(0)
        
    
    ################################################################################
    # Check loaded image
    if image is not None:
        
        # backup image
        orig = image.copy()

        # get the dimension of the image
        img_height, img_width, img_channels = orig.shape

        # get image file information 
        file_size = int(os.path.getsize(image_file)/MBFACTOR)
        
        print("Image file size: {} MB, dimension: {} X {}, channels : {}\n".format(str(file_size), img_height, img_width, img_channels))
        

        cleaned_thresh = orig
        
        masked_rgb_seg = orig
        
        # marker_detect(image)
        #(img_overlay, avg_width, pixel_cm_ratio) = marker_detect(image.copy())
        
        
        ######################################################################################

        # PhotoRoom Remove Background API
        # AI pre-trained model to segment plant object, return mask
        thresh_seg = remove(orig, only_mask = True).copy()
        

        #####################################################################################
        # find the largest contour in the threshold image
        
        cnts = cv2.findContours(thresh_seg.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        cnts = imutils.grab_contours(cnts)
        c = max(cnts, key=cv2.contourArea)
        
        
        blank_image = np.zeros(thresh_seg.shape, np.uint8)
        
        #cleaned_thresh = cv2.fillPoly(blank_image, pts = c, color = (255, 255, 255))
        
        cleaned_thresh = cv2.drawContours(blank_image, [c], -1, (255, 255, 255), cv2.FILLED)
        
        
        # Inverting the mask by 
        # performing bitwise-not operation 
        cleaned_thresh = cv2.bitwise_not(cleaned_thresh) 


        #cleaned_thresh = thresh_seg
        
        # use mask to generate segmentation object
        masked_rgb_seg = cv2.bitwise_and(orig, orig, mask = cleaned_thresh)
        

        # marker_detect(image)
        (img_overlay, avg_width, pixel_cm_ratio) = marker_detect(masked_rgb_seg.copy())
        

        '''
        ##############################################################################################
        n_cluster = 2
        
        args_channels = '0'
        
        args_colorspace = 'lab'
        
        thresh_cluster = color_cluster_seg(masked_rgb_seg, args_colorspace, args_channels, n_cluster)
        
        #thresh_cluster = cv2.threshold(thresh_cluster, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        
        
        masked_rgb_seg = cv2.bitwise_and(masked_rgb_seg, masked_rgb_seg, mask = thresh_cluster)
        
        #result_img_path = result_path + 'masked_rgb_seg.png'
        #cv2.imwrite(result_img_path, masked_rgb_seg)
        
        #result_img_path = result_path + 'thresh_seg.png'
        #cv2.imwrite(result_img_path, thresh_seg)
        
        
        #result_img_path = result_path + 'thresh.png'
        #cv2.imwrite(result_img_path, thresh)
        
        
        #masked_rgb_seg = cv2.drawContours(masked_rgb_seg, [c], -1, (0, 255, 0), 3)
        '''

    return cleaned_thresh, masked_rgb_seg, img_overlay, avg_width, pixel_cm_ratio
        


# get file information from the file path using python3
def get_file_info(file_full_path):
    
    p = pathlib.Path(file_full_path)

    filename = p.name

    basename = p.stem

    file_path = p.parent.absolute()

    file_path = os.path.join(file_path, '')
    
    return file_path, filename, basename



# save result files
def write_image_output(imagearray, result_path, base_name, addition, ext):

    # save segmentation result
    result_file = (result_path + base_name + addition + ext)
    
    #print(result_file)
    
    cv2.imwrite(result_file, imagearray)
    
    # check saved file
    if os.path.exists(result_file):
        print("Result file was saved at {0}\n".format(result_file))

    else:
        print("Result file writing failed!\n")
    




# save results as excel file
def write_excel_output(trait_file, trait_sum):
    
    if os.path.isfile(trait_file):
        # update values

        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'avg_width'
        sheet.cell(row = 1, column = 3).value = 'pixel_cm_ratio'

       
    for row in trait_sum:
        sheet.append(row)
   
   
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Result file was saved at {}\n".format(trait_file))
        
    else:
        print("Error in saving Result file\n")




# for Parallel processing
def batch_process(image_file):

    (file_path, filename, basename) = get_file_info(image_file)

    print("Segment foreground object for image file {} ...\n".format(file_path, filename, basename))

    # main pipeline to perform the segmentation based on u2net and color clustering
    (thresh, masked_rgb) = u2net_seg(image_file)

    # save mask result image as png format
    # write_image_output(thresh, result_path, basename, '_mask.', 'png')

    # save masked result image as png format
    write_image_output(masked_rgb, result_path, basename, '_masked.', 'png')







if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", dest = "path", type = str, required = True,    help = "path to image file")
    ap.add_argument("-ft", "--filetype", dest = "filetype", type = str, required = False, default='jpg,png', help = "Image filetype")
    ap.add_argument("-tp", "--temp_path", dest = "temp_path", type = str, required = True,  help = "template image path")
    ap.add_argument("-o", "--output_path", dest = "output_path", type = str, required = False,    help = "result path")
    ap.add_argument('-min', '--min_size', dest = "min_size", type = int, required = False, default = 1600,  help = 'min size of object to be segmented.')
    ap.add_argument('-max', '--max_size', dest = "max_size", type = int, required = False, default = 1000000,  help = 'max size of object to be segmented.')
    args = vars(ap.parse_args())
    

    # setup input and output file paths

    file_path = args["path"]
    ext = args['filetype']

    #accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype
    
    #accquire image file list
    imgList = sorted(glob.glob(image_file_path))

    
    # Load the template 
    template_file = args['temp_path'] + "template.jpg"
    
    template_path = os.path.abspath(template_file)

    
    if os.path.exists(template_path):
        
        print("Found template path {}\n".format(template_path))
        
        template = cv2.imread(template_path, 0)
        
    else:
        print("Error in loading template file\n")
    
    
    # result path
    mkpath = os.path.dirname(file_path) +'/detection'
    mkdir(mkpath)
    seg_path = mkpath + '/'
    
    
    # result file path
    result_path = args["output_path"] if args["output_path"] is not None else seg_path

    result_path = os.path.join(result_path, '')
    
    # print out result path
    print("results_folder: {}\n".format(result_path))

    '''
    #########################################################################
    # scan the folder to remove the 0 size image
    for image_id, image_file in enumerate(imgList):
        try:
            image = Image.open(image_file)
        except PIL.UnidentifiedImageError as e:
            print(f"Error in file {image_file}: {e}")
            os.remove(image_file)
            print(f"Removed file {image_file}")
    '''
    ############################################################################
    #accquire image file list after remove error images
    imgList = sorted(glob.glob(image_file_path))


    ########################################################################
    # parameters
    args_min_size = args['min_size']
    args_max_size = args['max_size']



    '''
    if args_parallel == 1:
        # Parallel processing
        #################################################################################
        import psutil
        from multiprocessing import Pool
        from contextlib import closing

        # parallel processing
        # get cpu number for parallel processing
        agents = psutil.cpu_count() - 2

        print("Using {0} cores to perform parallel processing... \n".format(int(agents)))

        # Create a pool of processes. By default, one is created for each CPU in the machine.
        with closing(Pool(processes=agents)) as pool:
            result = pool.map(batch_process, imgList)
            pool.terminate()

    else:
    '''
    #########################################################################
    # analysis pipeline
    # loop execute
    '''
    # marker result path
    mkpath = os.path.dirname(file_path) +'/cropped'
    mkdir(mkpath)
    marker_path = mkpath + '/'
    '''
    
    # save result as an excel file
    ratio_sum = []
    

    for image_id, image_file in enumerate(imgList):
        # store iteration start timestamp
        start = time.time()

        (file_path, filename, basename) = get_file_info(image_file)
        
        print("Plant object segmentation using u2net model for image {} ... \n".format(file_path))

        # main pipeline to perform the segmentation based on u2net and color clustering
        (cleaned_thresh, masked_rgb_seg, img_overlay, avg_width, pixel_cm_ratio) = u2net_seg(image_file)

        # save masked result image as png format
        write_image_output(img_overlay, result_path, basename, '_masked.', ext)

        # store iteration end timestamp
        end = time.time()

        # show time of execution per iteration
        #print(f"Segmentation finished for: {filename}\tTime taken: {(end - start) * 10 ** 3:.03f}s !\n")

        print("Segmentation finished for: {} in --- {} seconds ---\n".format(filename, (end - start)))
        
        ratio_sum.append([filename, avg_width, pixel_cm_ratio])
        
        '''
        ################################################################
        # save marker detection results
        result_file = (marker_path + basename + '_md.' + ext)
        
        print("Saving file '{} '...\n".format(result_file))
        
        cv2.imwrite(result_file, img_overlay)
        ################################################################
        '''
        


    ratio_sum_file = (result_path + 'unit.xlsx')

    write_excel_output(ratio_sum_file, ratio_sum)


    

    
