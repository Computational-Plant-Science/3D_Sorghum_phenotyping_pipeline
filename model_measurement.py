"""
Version: 1.5

Summary: compute the whole root traits based on 3D Sorghum model

Author: Suxing liu

Author-email: suxingliu@gmail.com

USAGE:

    python3 model_measurement.py -i ~/example/test.ply  -o ~/example/ --n_plane 10
    
INPUT:

    3D Sorghum model, aligned along Z direction in 3D coordinates.

OUTPUT:

    Excel file contains traits computation results

PARAMETERS:
    ("-i", "--input", dest="input", required=True, type=str, help="full path to 3D model file")
    ("-o", "--output_path", dest = "output_path", type = str, required = False, help = "result path")
    ("--n_plane", dest = "n_plane", type = int, required = False, default = 5,  help = "Number of planes to segment the 3d model along Z direction")
    ("--visualize", dest = "visualize", required = False, type = int, default = 0, help = "Display model or not, default not display")

"""
#!/usr/bin/env python


import glob
import numpy as np
import os
import sys
import pathlib
import argparse

import open3d as o3d
import copy

import openpyxl

import matplotlib
matplotlib.use('Agg')
from matplotlib import pyplot as plt
import math

# import warnings filter
from warnings import simplefilter
# ignore all future warnings
simplefilter(action='ignore', category=FutureWarning)




# generate folder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        #shutil.rmtree(path)
        #os.makedirs(path)
        return False









# compute the center coordinates of a 3d point cloud by slicing it into n_plane segments
def get_pt_sel_parameter(Data_array_pt, n_plane):
    
    ####################################################################
    
    # load skeleton coordinates and radius 
    Z_pt_sorted = np.sort(Data_array_pt[:,2])
    
    pt_plane = []
    
    
    # initialize paramters
    pt_plane_center = []
    
    pt_plane_diameter_max = []
    
    pt_plane_diameter_min = []
    
    pt_plane_diameter_avg = []
    

    
    filter_plane_center = []
    
    filter_plane_volume = []
    
    filter_plane_eccentricity = []
    
    filter_plane_bushiness = []
    
    
    # slicing models based number of planes along Z axis
    for idx, x in enumerate(range(n_plane)):
        
        ratio_s = idx/n_plane
        ratio_e = (idx+1)/n_plane
        
        print("ratio_s ratio_e {} {}\n".format(ratio_s, ratio_e))
        
        # index of end plane 
        idx_sel_e = int(len(Z_pt_sorted)*ratio_e) 
    
        Z_e = Z_pt_sorted[idx_sel_e]  if idx_sel_e < len(Data_array_pt) else (len(Data_array_pt) - 1)
        
        # index of start plane
        idx_sel_s = int(len(Z_pt_sorted)*ratio_s) 
    
        Z_s = Z_pt_sorted[idx_sel_s]

        # mask between the start and end plane
        Z_mask = (Data_array_pt[:,2] <= Z_e) & (Data_array_pt[:,2] >= Z_s) 
        
        Z_pt_sel = Data_array_pt[Z_mask]
        
        
        #print(Z_pt_sel.shape)
        
        # initialize the o3d object
        pcd_Z_mask = o3d.geometry.PointCloud()
    
        pcd_Z_mask.points = o3d.utility.Vector3dVector(Z_pt_sel)
        
        
        # get the diameter of the sliced model 
        (pt_diameter_max, pt_diameter_min, pt_diameter_avg, pt_length, pt_volume, pt_ob_volume) = get_pt_parameter(pcd_Z_mask)
        
        print("Current slice diameter_max = {}, diameter_min = {}, diameter_avg = {}\n".format(pt_diameter_max, pt_diameter_min, pt_diameter_avg))
        
        # get the model center position
        model_center = pcd_Z_mask.get_center()

        pt_plane.append(pcd_Z_mask)
        
        pt_plane_center.append(model_center)
        
        #pt_plane_diameter.append(pt_diameter)
        
        pt_plane_diameter_max.append(pt_diameter_max)
        
        pt_plane_diameter_min.append(pt_diameter_min)
        
        pt_plane_diameter_avg.append(pt_diameter_avg)
        
        filter_plane_bushiness.append(pt_volume/pt_ob_volume)
        
        # filter sliced models using sphere with radius and compute parameters
        ################################################################
        # copy current sliced model
        pt_sel_filter = copy.deepcopy(pcd_Z_mask)
        
        # get 3d points
        points = np.asarray(pt_sel_filter.points)

        # Sphere center and radius
        radius = pt_diameter_avg*0.5
        
        print("radius =  {} \n".format(radius))

        # Calculate distances to center, set new points
        distances = np.linalg.norm(points - model_center, axis=1)
        
        pt_sel_filter.points = o3d.utility.Vector3dVector(points[distances <= radius])
        
        # filter sliced model
        (filter_diameter_max, filter_diameter_min, filter_diameter, filter_length, filter_volume, filter_density) = get_pt_parameter(pcd_Z_mask)
        
        
        filter_plane_center.append(pt_sel_filter.get_center())
        
        filter_plane_volume.append(filter_volume)
                
        #########################################################################
        # compute eccentricity using oriented bounding box axis
        
        # get OrientedBoundingBox
        obb = pt_sel_filter.get_oriented_bounding_box()

        # assign color for OrientedBoundingBox
        obb.color = (0, 0, 1)

        # get the eight points that define the bounding box.
        pcd_coord = obb.get_box_points()

        #print(obb.get_box_points())

        #pcd_coord.color = (1, 0, 0)

        # From Open3D to numpy array
        np_points = np.asarray(pcd_coord)

        # create Open3D format for points 
        #pcd_coord = o3d.geometry.PointCloud()
        #pcd_coord.points = o3d.utility.Vector3dVector(np_points)
    
        # check the length of the joint 3 vector in the bounding box to estimate the orientation of model
        list_dis = [np.linalg.norm(np_points[0] - np_points[1]), np.linalg.norm(np_points[0] - np_points[2]), np.linalg.norm(np_points[0] - np_points[3])]
        
        #print("list_dis =  {} \n".format(list_dis))
        
        filter_plane_eccentricity.append(min(list_dis[0],list_dis[1])/max(list_dis[0],list_dis[1]))
        
        #print("filter_plane_eccentricity =  {} \n".format(filter_plane_eccentricity))
        
        # get rotation matrix
        #rotation_array = obb.R.tolist()
        # get the eight points that define the bounding box.
        #pcd_coord = obb.get_box_points()
        #print("obb.R =  {} \n".format(obb.R))
        #rotation_array = obb.R.tolist()
        #r = R.from_matrix(rotation_array)
        #orientation_angle = r.as_euler('xyz', degrees=True)
        #print("orientation_angle =  {} \n".format(orientation_angle))
        
        

        
        ################################################################
        
        #pt_plane_volume.append(pt_volume)
        

    return pt_plane, pt_plane_center, pt_plane_diameter_max, pt_plane_diameter_min, pt_plane_diameter_avg, filter_plane_center, filter_plane_volume, filter_plane_eccentricity, filter_plane_bushiness
    

    

# compute dimensions of point cloud
def get_pt_parameter(pcd):
    
    # get convex hull of a point cloud is the smallest convex set that contains all points.
    hull, _ = pcd.compute_convex_hull()
    #hull_ls = o3d.geometry.LineSet.create_from_triangle_mesh(hull)
    #hull_ls.paint_uniform_color((1, 0, 0))
    
    # get AxisAlignedBoundingBox
    aabb = pcd.get_axis_aligned_bounding_box()
    aabb.color = (0, 1, 0)
    
    #Get the extent/length of the bounding box in x, y, and z dimension.
    aabb_extent = aabb.get_extent()
    
    aabb_extent_half = aabb.get_half_extent()
    
    # get OrientedBoundingBox
    obb = pcd.get_oriented_bounding_box()
    
    obb.color = (0, 0, 1)

    #visualize the convex hull as a red LineSet
    #o3d.visualization.draw_geometries([pcd, aabb, obb, hull_ls])
    
    
    # compute parameters
    #pt_diameter_max = max(aabb_extent[0], aabb_extent[1])
    
    pt_diameter_max = (math.sqrt(pow(aabb_extent[0],2) + pow(aabb_extent[1],2)) + max(aabb_extent[0], aabb_extent[1])) / 2.0
    
    pt_diameter_min = min(aabb_extent_half[0], aabb_extent_half[1])
    
    pt_diameter_avg = (pt_diameter_max + pt_diameter_min)*0.5

    pt_length = (aabb_extent[2])

    # compute as cylinder
    #pt_volume = np.pi * ((pt_diameter*0.5) ** 2) * pt_length
    
    # compute as convexhull volume
    pt_volume = hull.get_volume()
    
    # oriented bounding box volume
    pt_ob_volume = pcd.get_oriented_bounding_box().volume()

    return pt_diameter_max, pt_diameter_min, pt_diameter_avg, pt_length, pt_volume, pt_ob_volume
    
    
    
    

#colormap mapping
def get_cmap(n, name = 'hsv'):
    """get the color mapping""" 
    #viridis, BrBG, hsv, copper, Spectral
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    return plt.cm.get_cmap(name,n+1)




# save point cloud data from numpy array as ply file, compatible format with open3d library
def write_ply(path, data_numpy_array):
    
    #data_range = 100
    
    #Normalize data range for generate cross section level set scan
    #min_max_scaler = preprocessing.MinMaxScaler(feature_range = (0, data_range))

    #point_normalized = min_max_scaler.fit_transform(data_numpy_array)
    
    #initialize pcd object for open3d 
    pcd = o3d.geometry.PointCloud()
     
    pcd.points = o3d.utility.Vector3dVector(data_numpy_array)
    
    # get the model center position
    model_center = pcd.get_center()
    
    # geometry points are translated directly to the model_center position
    pcd.translate(-1*(model_center))
    
    #write out point cloud file
    o3d.io.write_point_cloud(path, pcd, write_ascii = True)
    
    
    # check saved file
    if os.path.exists(path):
        print("Converted 3d model was saved at {0}".format(path))
        return True
    else:
        return False
        print("Model file converter failed !")
        #sys.exit(0)


# compute diameter from area
def area_radius(area_of_circle):
    radius = ((area_of_circle/ math.pi)** 0.5)
    
    #note: return diameter instead of radius
    return 2*radius 



#compute angle
def angle(directions):
    """Return the angle between vectors"""
    vec2 = directions[1:]
    vec1 = directions[:-1]

    norm1 = np.sqrt((vec1 ** 2).sum(axis=1))
    norm2 = np.sqrt((vec2 ** 2).sum(axis=1))
    cos = (vec1 * vec2).sum(axis=1) / (norm1 * norm2)   
    return np.arccos(cos)


#first derivative function
def first_derivative(x) :

    return x[2:] - x[0:-2]


#second derivative function
def second_derivative(x) :
    
    return x[2:] - 2 * x[1:-1] + x[:-2]


#compute curvature
def curvature(x, y) :

    x_1 = first_derivative(x)
    x_2 = second_derivative(x)
    y_1 = first_derivative(y)
    y_2 = second_derivative(y)
    return np.abs(x_1 * y_2 - y_1 * x_2) / np.sqrt((x_1**2 + y_1**2)**3)


#compute angle between two vectors(works for n-dimensional vector),
def dot_product_angle(v1,v2):

    if np.linalg.norm(v1) == 0 or np.linalg.norm(v2) == 0:
        
        print("Zero magnitude vector!")
        
        return 0
        
    else:
        vector_dot_product = np.dot(v1,v2)
        
        arccos = np.arccos(vector_dot_product / (np.linalg.norm(v1) * np.linalg.norm(v2)))
        
        angle = np.degrees(arccos)
        
        return (90 - angle)
        
    '''
    if angle > 0 and angle < 45:
        return (90 - angle)
    elif angle < 90:
        return angle
    else:
        return (180- angle)
    '''


# Traits analysis for the input 3D model
def analyze_pt(pt_file):
    
    ###################################################################
    #load aligned ply point cloud file
        
    print("Loading 3D point cloud {}...\n".format(pt_file))

    pcd = o3d.io.read_point_cloud(pt_file)
    
    Data_array_pcloud = np.asarray(pcd.points)
    
    print(Data_array_pcloud.shape)
    

    # load skeleton coordinates and radius 
    X = Data_array_pcloud[:,0] 
    Y = Data_array_pcloud[:,1] 
    Z = Data_array_pcloud[:,2] 


    #print(X.shape, Y.shape, Z.shape)
    

    
    #for idx, x in enumerate(range(n_plane)):

    if pcd.has_colors():
        
        print("Render colored point cloud\n")
        
        pcd_color = np.asarray(pcd.colors)
        
        if len(pcd_color) > 0: 
            
            pcd_color = np.rint(pcd_color * 255.0)
        
        #pcd_color = tuple(map(tuple, pcd_color))
    else:
        
        print("Generate random color\n")
    
        pcd_color = np.random.randint(256, size = (len(Data_array_pcloud),3))
    

    #initialize parameters
    pt_diameter_max = pt_diameter_min = pt_length = pt_diameter = pt_eccentricity = pt_density = pt_angle = pt_angle_max = pt_angle_min = sum_volume = 0
    
    
    #Eccentricity
    
    #Bushiness

    #compute dimensions of point cloud data
    (pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_volume, pt_ob_volume) = get_pt_parameter(pcd)

    print("pt_diameter_max = {} pt_diameter_min = {} pt_diameter = {} pt_length = {} pt_volume = {}\n".format(pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_volume))
    
    
    
    ########################################################################################################3
    # slicing models using n_plane
    print("Using {} planes to scan the model along Z axis...".format(n_plane))
    
     
    (pt_plane, pt_plane_center, pt_plane_diameter_max, pt_plane_diameter_min, pt_plane_diameter_avg, filter_plane_center, filter_plane_volume, filter_plane_eccentricity, filter_plane_bushiness) = get_pt_sel_parameter(Data_array_pcloud, n_plane)
    
    #o3d.visualization.draw_geometries(pt_plane)
    
    pt_center_arr = np.vstack(pt_plane_center)
    
    # compute simplified center vector angles
    # construct vectors
    #start_v = [pt_plane_center[0][0], pt_plane_center[0][1], pt_plane_center[0][2]]
    #end_v = [pt_plane_center[0][0] - pt_plane_center[-1][0], pt_plane_center[0][1] - pt_plane_center[-1][1], pt_plane_center[0][2] - pt_plane_center[-1][2]]
    #pt_angle = dot_product_angle(start_v, end_v)
    #print(pt_angle)
    
    
    
    # initialize parameters
    filter_plane_angle = []
    
    
    # define unit vector
    v_x = [1,0,0]
    v_y = [0,1,0]
    v_z = [0,0,1]


    # compute side angles for each sliced model
    for idx, f_center in enumerate(filter_plane_center):
        
        if idx > 0:
            
            #print(idx, f_center)
            
            center_vector = [f_center[0] - filter_plane_center[idx-1][0], f_center[1] - filter_plane_center[idx-1][1], f_center[2] - filter_plane_center[idx-1][2]]
        
            norm_center_vector = center_vector / np.linalg.norm(center_vector)
        
            cur_angle = dot_product_angle(norm_center_vector, v_z)
            
            #print("cur_angle = {} ...".format(cur_angle))
    
            filter_plane_angle.append(cur_angle)
    
    
    pt_angle = np.mean(filter_plane_angle)
    
    pt_angle_max = max(filter_plane_angle)
    
    pt_angle_min = min(filter_plane_angle)
    
    print("pt_angle = {}, pt_angle_max = {}, pt_angle_min = {}\n".format(pt_angle, pt_angle_max, pt_angle_min))
    

    # Sum of all volume for each sliced model 
    sum_volume = sum(filter_plane_volume)

    # average of eccentricity
    avg_eccentricity = np.mean(filter_plane_eccentricity)
    
    # average of bushiness
    avg_bushiness = np.mean(filter_plane_bushiness)

    #Visualization pipeline
    ####################################################################
    # The number of points per line
    
    if visualize == 1:
    
        '''
        from mayavi import mlab
        from tvtk.api import tvtk
        
        mlab.figure("point_cloud", size = (800, 800), bgcolor = (0, 0, 0))
        
        mlab.clf()


        #visualize point cloud model with color
        ####################################################################

        x, y, z = Data_array_pcloud[:,0], Data_array_pcloud[:,1], Data_array_pcloud[:,2] 
        
        
        pts = mlab.points3d(x,y,z, mode = 'point')
        
        sc = tvtk.UnsignedCharArray()
        
        sc.from_array(pcd_color)

        pts.mlab_source.dataset.point_data.scalars = sc
        
        pts.mlab_source.dataset.modified()

        #####################################################################
        #visualize point cloud model with color
        pts = mlab.points3d(pt_center_arr[:,0], pt_center_arr[:,1], pt_center_arr[:,2], color = (1,0,0), mode = 'sphere', scale_factor = 0.08)
        
        pts = mlab.plot3d(pt_center_arr[:,0], pt_center_arr[:,1], pt_center_arr[:,2], tube_radius = 0.025, color = (0,1,0))

        mlab.show()
        '''
    
        # Visualization of center curve
        ####################################################################
        filter_center_points = np.vstack(filter_plane_center)
        
        filter_center_line = []
        
        for i in range(n_plane):
            
            if i+1 < n_plane:
                filter_center_line.append([i, i+1])

        
        colors_filter = [[1, 0, 0] for i in range(n_plane-1)]

        
        lines_filter_set = o3d.geometry.LineSet()
        lines_filter_set.points = o3d.utility.Vector3dVector(filter_center_points)
        lines_filter_set.lines = o3d.utility.Vector2iVector(filter_center_line)
        lines_filter_set.colors = o3d.utility.Vector3dVector(colors_filter)
        
        
        vis = o3d.visualization.Visualizer()
        vis.create_window()
        vis.add_geometry(lines_filter_set)
        vis.add_geometry(pcd)
        vis.get_render_option().line_width = 5
        vis.get_render_option().point_size = 1
        vis.run()



    return pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_angle, pt_angle_max, pt_angle_min, sum_volume, avg_eccentricity, avg_bushiness




# get file information from the file path using python3
def get_file_info(file_full_path):
    
    p = pathlib.Path(file_full_path)
    filename = p.name
    basename = p.stem

    file_path = p.parent.absolute()
    file_path = os.path.join(file_path, '')

    return file_path, filename, basename



# save results as excel file
def write_output(trait_file, trait_sum):
    
    if os.path.isfile(trait_file):
        # update values

        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'root system diameter max'
        sheet.cell(row = 1, column = 2).value = 'root system diameter min'
        sheet.cell(row = 1, column = 3).value = 'root system diameter'
        sheet.cell(row = 1, column = 4).value = 'root system length'
        sheet.cell(row = 1, column = 5).value = 'root system angle'
        sheet.cell(row = 1, column = 6).value = 'root system angle max'
        sheet.cell(row = 1, column = 7).value = 'root system angle min'
        sheet.cell(row = 1, column = 8).value = 'root system volume'
        sheet.cell(row = 1, column = 9).value = 'root system eccentricity'
        sheet.cell(row = 1, column = 10).value = 'root system bushiness'

       
    for row in trait_sum:
        sheet.append(row)
   
   
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Result file was saved at {}\n".format(trait_file))
        
    else:
        print("Error in saving Result file\n")




if __name__ == '__main__':
    

    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-i", "--input", dest="input", required=True, type=str, help="full path to 3D model file")
    #ap.add_argument("-p", "--path", dest = "path", required = True, type = str, help = "path to 3D model file")
    #ap.add_argument("-ft", "--filetype", dest = "filetype", type = str, required = False, default = 'ply', help = "3D model file filetype, default *.ply")
    ap.add_argument("-o", "--output_path", dest = "output_path", type = str, required = False, help = "result path")
    ap.add_argument("--n_plane", dest = "n_plane", type = int, required = False, default = 10,  help = "Number of planes to segment the 3d model along Z direction")
    ap.add_argument("--visualize", dest = "visualize", required = False, type = int, default = 0, help = "Display model or not, default not display")
    args = vars(ap.parse_args())

    


    
    if os.path.isfile(args["input"]):

        input_file = args["input"]

        (file_path, filename, basename) = get_file_info(input_file)

        print("Processing 3d model point cloud file '{} {} {}'...\n".format(file_path, filename, basename))

        # result path
        result_path = args["output_path"] if args["output_path"] is not None else file_path

        result_path = os.path.join(result_path, '')

        # print out result path
        print("results_folder: {}\n".format(result_path))

        # number of slices for cross section
        n_plane = args['n_plane']

        visualize = args["visualize"]

        # start pipeline
        ########################################################################################3
        # compute parameters
        (pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_angle, pt_angle_max, pt_angle_min, pt_volume, avg_eccentricity, avg_bushiness) = analyze_pt(input_file)
        
        # save result as an excel file
        trait_sum = []

        trait_sum.append([pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_angle, pt_angle_max, pt_angle_min, pt_volume, avg_eccentricity, avg_bushiness])

        trait_file = (result_path + basename + '_trait.xlsx')

        write_output(trait_file, trait_sum)


    else:

        print("The input file is missing or not readable!\n")

        print("Exiting the program...")

        sys.exit(0)


    
    '''
    # loop multiple files for bath processing 
    # python3 model_measurement.py -p ~/example/ -ft ply -o ~/example/ -n 5 -v 0
    #######################################################################################
    
    # path to model file 
    file_path = args["path"]
    
    ext = args['filetype']
    
    files = file_path + '*.' + ext
    
    n_plane = args['n_plane']

    visualize = args["visualize"]
    
    
    # obtain image file list
    fileList = sorted(glob.glob(files))


    

    for input_file in fileList:
        
        if os.path.isfile(input_file):
            
            (file_path, filename, basename) = get_file_info(input_file)

            print("Processing 3d model point cloud file '{} {} {}'...\n".format(file_path, filename, basename))

            # result path
            result_path = args["output_path"] if args["output_path"] is not None else file_path

            result_path = os.path.join(result_path, '')

            # print out result path
            print("results_folder: {}\n".format(result_path))

            # number of slices for cross section
            n_plane = args['n_plane']

            visualize = args["visualize"]

            # start pipeline
            ########################################################################################3
            # compute parameters
            (s_diameter_max, s_diameter_min, s_diameter, s_length, avg_density, avg_volume) = analyze_pt(input_file)

            # save result as an excel file
            trait_sum = []

            trait_sum.append([s_diameter_max, s_diameter_min, s_diameter, s_length, avg_density, avg_volume])

            trait_file = (result_path + basename + '_trait.xlsx')

            write_output(trait_file, trait_sum)
                


        else:
        
            print("The input file is missing or not readable!\n")
            
            print("Exiting the program...")
            
            sys.exit(0)
    '''


    
