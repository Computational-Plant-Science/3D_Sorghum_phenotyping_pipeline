'''
Name: marker_detection.py

Version: 1.0

Summary: Detect sticker shape markers in image and cropp image based on marker location
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2024-09-09

USAGE:

    python3 marker_detection.py -p ~/example/portable_scanner/bean_test/30_stop_220/ -ft jpg



'''

# import necessary packages
import argparse
import cv2
import numpy as np
import os
import glob
#from pathlib import Path 
import pathlib
import shutil 


from tabulate import tabulate
import openpyxl

#from pathlib import Path


# generate folder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        shutil.rmtree(path)
        os.makedirs(path)
        return False



# get file information from the file path using python3
def get_file_info(file_full_path):
    
    p = pathlib.Path(file_full_path)
    filename = p.name
    basename = p.stem

    file_path = p.parent.absolute()
    file_path = os.path.join(file_path, '')

    return file_path, filename, basename




# save results as excel file
def write_output(trait_file, trait_sum):
    
    if os.path.isfile(trait_file):
        # update values

        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'avg_width'
        sheet.cell(row = 1, column = 3).value = 'pixel_cm_ratio'

       
    for row in trait_sum:
        sheet.append(row)
   
   
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Result file was saved at {}\n".format(trait_file))
        
    else:
        print("Error in saving Result file\n")




# Detect stickers in the image
def marker_detect(image_file):
    

    # load the image, clone it for output, and then convert it to grayscale
    img_ori = cv2.imread(image_file)
    
    img_rgb = img_ori.copy()
      
    # convert the input image to a grayscale
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY) 
      
    # Store width and height of template in w and h 
    #w, h = template.shape[::-1] 
    
    # get the dimension of the image
    img_height, img_width, img_channels = img_ori.shape

    img_overlay = img_rgb
    
    (ret, thresh) = cv2.threshold(img_gray, 50, 255, 0)
    
    # Find the contours in the image using cv2.findContours() function.
    contours,hierarchy = cv2.findContours(thresh, 1, 2)
    
    print("Number of contours detected:", len(contours))
    
    
    i = 0
    
    # initialize square width 
    width_rec = []

    
    # list for storing names of shapes 
    for cnt in contours:
        
        # here we are ignoring first counter because  
        # findcontour function detects whole image as shape 
        if i == 0: 
            i = 1
            continue

        x1,y1 = cnt[0][0]
        
              
        # cv2.approxPloyDP() function to approximate the shape
        approx = cv2.approxPolyDP(cnt, 0.01 * cv2.arcLength(cnt, True), True)

        if len(approx) == 4:
          
            (x, y, w, h) = cv2.boundingRect(cnt)
            
            # compute the center of the contour
            M = cv2.moments(cnt)
            
            if M["m00"] != 0:
                cX = int(M["m10"] / M["m00"])
                cY = int(M["m01"] / M["m00"])
            else:
                # set values as what you need in the situation
                cX, cY = 0, 0

            ratio = float(w)/h
            
            if (cX < img_width*0.3 or cX > img_width*0.6) and (cY > img_width*0.3 or cY < img_width*0.6):
                
                    # define threshold for the dimension of square 
                    if min(w,h) > 80 and max(w,h) < 300:
                        
                        if ratio >= 0.7 and ratio <= 1.2:

                            img_overlay = cv2.drawContours(img_rgb, [cnt], -1, (0,255,255), 5)

                            width_rec.append((w+h)*0.5)

    # compute the average of detected square dimension in pixels
    if len(width_rec) > 0:
        
        avg_width = np.mean(width_rec)
        
        pixel_cm_ratio = avg_width/2.5
     
    else:
        
        avg_width = 0
        pixel_cm_ratio = 0
    
    return img_overlay, avg_width, pixel_cm_ratio
    
    





if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help = "path to image file")
    ap.add_argument("-ft", "--filetype", required=True,    help = "image filetype")
    ap.add_argument("-o", "--output_path", dest = "output_path", type = str, required = False, help = "result path")
    args = vars(ap.parse_args())
    
    
    # Setting path to image files
    file_path = args["path"]
    ext = args['filetype']

    # Extract file type and path
    filetype = '*.' + ext
    image_file_path = file_path + filetype
    
    # Accquire image file list
    imgList = sorted(glob.glob(image_file_path))
    
    

    #imgList = (glob.glob(image_file_path))

    #print((imgList))
    #global save_path
    
    # Get number of images in the data folder
    n_images = len(imgList)
    
    result_list = []
    
    
    # save folder construction
    mkpath = os.path.dirname(file_path) +'/cropped'
    mkdir(mkpath)
    result_path = mkpath + '/'

    # print out result path
    print("results_folder: {}\n".format(result_path))
    
    
    # save result as an excel file
    trait_sum = []
    
    # Loop execute
    for image_file in imgList:
        
        #(image_file_name, sticker_overlay) = marker_detect(image)
        
        #result_list.append([image_file_name, sticker_overlay])
        
        (file_path, filename, basename) = get_file_info(image_file)

        print("Processing file '{} {} {}'...\n".format(file_path, filename, basename))
        
        
        #marker_detect(image)
        (img_overlay, avg_width, pixel_cm_ratio) = marker_detect(image_file)
        
        result_file = (result_path + basename + '_md.' + ext)
        
        print("Saving file '{} '...\n".format(result_file))
        
        cv2.imwrite(result_file, img_overlay)


        trait_sum.append([filename, avg_width, pixel_cm_ratio])

    
    
    trait_file = (result_path + 'trait.xlsx')

    write_output(trait_file, trait_sum)

